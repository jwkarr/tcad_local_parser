from openpyxl import load_workbook
from pathlib import Path

excel_file = Path("temp_extract/Legacy8.0.30-AppraisalExportLayout.xlsx")

wb = load_workbook(excel_file, data_only=True)

# Get the Property sheet
if "Property" in wb.sheetnames:
    ws = wb["Property"]
    print(f"Reading Property sheet: {ws.title}")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}\n")
    
    # Find the header row (should contain 'Field Name', 'Start', 'End')
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and row[0] and 'Field Name' in str(row[0]):
            header_row = i
            print(f"Header found at row {i}: {row}")
            break
    
    if header_row:
        print(f"\nField definitions:")
        print("-" * 100)
        
        # Read field definitions
        field_specs = {}
        for i, row in enumerate(ws.iter_rows(values_only=True, min_row=header_row+1), start=header_row+1):
            if not row or not row[0] or row[0] is None:
                break
            
            field_name = str(row[0]).strip()
            if not field_name:
                break
                
            start = row[2] if len(row) > 2 and row[2] is not None else None
            end = row[3] if len(row) > 3 and row[3] is not None else None
            length = row[4] if len(row) > 4 and row[4] is not None else None
            desc = row[5] if len(row) > 5 and row[5] is not None else ""
            
            if start and end:
                print(f"{i:4d}. {field_name:30s} Start: {start:5d} End: {end:5d} Len: {length} - {desc}")
                field_specs[field_name] = (start, end)
        
        print(f"\n\n=== FIELD_SPECS dictionary format ===")
        print("FIELD_SPECS = {")
        for field_name, (start, end) in sorted(field_specs.items(), key=lambda x: x[1][0]):
            print(f'    "{field_name}": ({start}, {end}),')
        print("}")
        
else:
    print("Property sheet not found!")
    print(f"Available sheets: {wb.sheetnames}")



