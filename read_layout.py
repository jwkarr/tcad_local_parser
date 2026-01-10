from openpyxl import load_workbook
from pathlib import Path

excel_file = Path("temp_extract/Legacy8.0.30-AppraisalExportLayout.xlsx")

try:
    wb = load_workbook(excel_file, data_only=True)
    print(f"Sheet names: {wb.sheetnames}")
    
    # Try the first sheet
    ws = wb.active
    print(f"\nActive sheet: {ws.title}")
    print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
    
    # Print first 50 rows to see the structure
    print("\nFirst 50 rows of data:")
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > 50:
            break
        print(f"Row {i}: {row}")
        
except Exception as e:
    print(f"Error reading Excel file: {e}")
    import traceback
    traceback.print_exc()



